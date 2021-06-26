import openpyxl

book = openpyxl.load_workbook("F:\\selenium\\PythonDemo.xlsx")
sheet = book.active
Dict1 = {}

cell = sheet.cell(row=1, column=2)
print(cell.value)
# sheet.cell(row=4, column=4).value = "Awesome"  # Currently this is not working
print(sheet.cell(row=4, column=4))

print(sheet.max_row)
print(sheet.max_column)
print(sheet['A5'].value)

for i in range(1, sheet.max_row+1):              # to get rows
    if sheet.cell(row=i, column=1).value == "TestCase2":      # to get value for only one testcase/entry
        for j in range(2, sheet.max_column+1):         # to get columns  , starting from 2 to get only values
            # print(sheet.cell(row=i, column=j).value)
            # Dict1["email="]="...."
            Dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

print(Dict1)
