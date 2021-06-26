import openpyxl


# If we want to call a method from a class.method name, it should be static method. for all others, new need to create objects
# self is also not required in it then to declare.
class HomePageData:
    test_HomePage_Data = [{"firstname": "Rahul", "email": "abc@yahoo.in", "gender": "Male"},
                          {"firstname": "Divya", "email": "xyz@yahoo.in", "gender": "Female"}]

    @staticmethod          # this is all to get data from excel. Not used in our code.
    def testDataFromExcel(test_case_name):
        Dict1 = {}
        book = openpyxl.load_workbook("F:\\selenium\\PythonDemo.xlsx")
        sheet = book.active
        Dict1 = {}
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:  # to get value for only one testcase/entry
                for j in range(2, sheet.max_column + 1):  # to get columns  , starting from 2 to get only values
                    # print(sheet.cell(row=i, column=j).value)
                    # Dict1["email="]="...."
                    Dict1[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        # print(Dict1)
        return [Dict1]
